import sys
import csv
from pathlib import Path

from openpyxl import load_workbook


def convert_xlsx_to_csv(xlsx_path: str, csv_path: str) -> None:
    x_path = Path(xlsx_path)
    c_path = Path(csv_path)

    wb = load_workbook(filename=str(x_path), read_only=True, data_only=True)
    ws = wb.active

    with c_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])


def main(argv: list[str]) -> int:
    if len(argv) != 3:
        sys.stderr.write("Usage: xlsx_to_csv.py <input.xlsx> <output.csv>\n")
        return 1

    xlsx_path, csv_path = argv[1], argv[2]

    try:
        convert_xlsx_to_csv(xlsx_path, csv_path)
    except Exception as exc:  # noqa: BLE001
        sys.stderr.write(f"ERROR: {exc}\n")
        return 1

    sys.stdout.write("OK\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
